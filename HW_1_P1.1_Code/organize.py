import pandas as pd
import matplotlib.pyplot as plt
import subprocess
import re
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import matplotlib.dates as mdates
import nltk
from nltk.corpus import words
import ssl
from collections import Counter

#run from root, file split into two parts, put everything from dump into a csv file,
# then run analysis on csv file, and finally plot the graph in a seperate tab. So tab with all csv information,
# tab with analysis and then tab with graphs
#REMEMBER TO SET JAVA HOME

CRAWL_PATH = "crawl_old/crawldb"
DUMP_FILE = "dump/part-r-00000"
OUTPUT = "output/output.xlsx"
LOG_PATH = "logs/2026-05/hadoop_fin.log"
OUTPUT_IMAGE = "output/output_graph.png"
SEG_PATH = "/Users/isabelledarling/Downloads/apache-nutch-1.222/dump_seg/dump"

#yeah idk why its doing this for nltk but this works to fix it by creating http context
try:
    _create_unverified_https_context = ssl._create_unverified_context
except AttributeError:
    pass
else:
    ssl._create_default_https_context = _create_unverified_https_context

nltk.download('words', quiet=True)
english_vocab = set(w.lower() for w in words.words())

#HELPER
def get_keyword_counts(SEG_PATH):
    #take out mostly html words
    stopwords = {
        # nutch
        'fetch', 'time', 'status', 'version', '2026', 'href', 'null', 'https', 'http',
        'metadata', 'score', 'retries', 'linked', 'anchor', 'signature', 'content', 'type',
        'crawldatum', 'modified', 'days', 'outlink', 'inlink', 'parse', 'parser', 'parsetext',
        'charset', 'protocol', 'host', 'domain', 'date', 'data',

        # HTML
        'class', 'font', 'menu', 'span', 'width', 'height', 'flex', 'field', 'title',
        'div', 'style', 'margin', 'padding', 'color', 'background', 'align', 'size',
        'text', 'html', 'body', 'script', 'javascript', 'true', 'false', 'image',
        'display', 'border', 'block', 'inline', 'hidden', 'main', 'header', 'footer',
        'nav', 'button', 'link', 'meta', 'hover', 'active', 'focus', 'rgba', 'item',
        'testid', 'dropdown', 'left', 'position', 'right', 'container', 'default',
        'center', 'level', 'none', 'webkit', 'solid', 'transparent', 'important',
        'absolute', 'relative', 'fixed', 'sticky', 'justify', 'overflow', 'scroll',
        'auto', 'line', 'weight', 'decoration', 'transform', 'transition', 'opacity',
        'index', 'zindex', 'wrapper', 'layout', 'grid', 'table', 'tbody', 'form',
        'input', 'submit', 'search', 'menuitem', 'checkbox', 'radio', 'dialog',
        'modal', 'icon', 'logo', 'aria', 'label', 'role', 'svg', 'path', 'stroke',

        # blah
        'cookie', 'privacy', 'policy', 'terms', 'conditions', 'copyright', 'rights',
        'reserved', 'login', 'signup', 'register', 'password', 'email', 'username',
        'click', 'here', 'read', 'more', 'loading', 'error', 'page', 'home', 'about',
        'contact', 'menu', 'skip', 'close', 'open', 'share', 'print',

        # eng
        'the', 'a', 'an', 'and', 'or', 'to', 'of', 'in', 'is', 'it', 'for', 'with',
        'on', 'as', 'by', 'at', 'this', 'that', 'from', 'are', 'was', 'were', 'be',
        'since', 'have', 'not', 'you', 'but', 'his', 'they', 'she', 'which', 'what',
        'their', 'will', 'there', 'been', 'much', 'many', 'would', 'could', 'should',
        'some', 'these', 'your', 'when', 'into', 'then', 'than', 'only', 'also'
    }

    stats_by_url = {}
    global_counter = Counter()

    current_url = None
    text_buffer = []
    in_parse_text = False

    with open(SEG_PATH, 'r', errors='ignore') as f:
        for line in f:
            line_stripped = line.strip()
            url_match = re.search(r"^(?:URL::\s*)?(https?://[^\s]+)", line_stripped)

            if url_match:
                if current_url and text_buffer:
                    content = " ".join(text_buffer)
                    words_list = re.findall(r'[a-zA-Z]+', content.lower())
                    filtered = [w for w in words_list if w not in stopwords and len(w) > 3]

                    if current_url not in stats_by_url:
                        stats_by_url[current_url] = set()
                    stats_by_url[current_url].update(filtered)
                    global_counter.update(filtered)

                current_url = url_match.group(1).strip().rstrip('/')

                if current_url not in stats_by_url:
                    stats_by_url[current_url] = set()

                text_buffer = []
                in_parse_text = False
                continue

            if "ParseText::" in line_stripped:
                in_parse_text = True
                continue

            if in_parse_text and re.match(r"^(?:Recno|URL|ParseData|ParseText|Content|CrawlDatum)::", line_stripped, re.IGNORECASE):
                in_parse_text = False

            if in_parse_text and current_url:
                text_buffer.append(line_stripped)

        if current_url and text_buffer:
            content = " ".join(text_buffer)
            words_list = re.findall(r'[a-zA-Z]+', content.lower())
            filtered = [w for w in words_list if w not in stopwords and len(w) > 3]

            if current_url not in stats_by_url:
                stats_by_url[current_url] = set()
            stats_by_url[current_url].update(filtered)
            global_counter.update(filtered)

    final_stats = {u: len(words) for u, words in stats_by_url.items()}
    return final_stats, global_counter

#list of what we are looking for, time, status, url, fetch time, score? Can add more if needed
#s for whitespace d for digit, basically telling the program to look for these patterns
url_pattern = re.compile(r"^(https?://\S+)")
status_pattern = re.compile(r"^Status:\s+(.+)")
score_pattern = re.compile(r"^Score:\s+([^\s]+)")
time_pattern = re.compile(r"^Fetch time:\s+(.+)")

rows = []
curr = {}

with open(DUMP_FILE, "r", errors="ignore", encoding="utf-8") as file:
    for line in file:
        line = line.strip()

        #URL
        url_obj = url_pattern.match(line)
        if url_obj:
            if curr and "URL" in curr:
                rows.append(curr)
            curr = {"URL" : url_obj.group(1)}
            continue

        if not curr:
            continue

        #STATUS
        status_obj = status_pattern.match(line)
        if status_obj:
            curr["Status"] = status_obj.group(1)
            continue

        #SCORE
        score_obj = score_pattern.match(line)
        if score_obj:
            curr["Score"] = score_obj.group(1)
            continue

        #TIME (find the time elasped in the pandas section)
        time_obj = time_pattern.match(line)
        if time_obj:
            curr["Fetch Time"] = time_obj.group(1)

if curr and "URL" in curr:
    rows.append(curr)

df = pd.DataFrame(rows)

#PANDAS, take csv info, make a new tab for anaylsis.
#Crawl speed🡪 #pages/minute, ratio of #URL crawled / #URL to be crawled, etc.)

#grab main stats
try:
    result = subprocess.run(["bin/nutch", "readdb", CRAWL_PATH, "-stats"], capture_output=True, text=True)
    out = result.stdout
except subprocess.CalledProcessError as e:
    exit()

#can change to grab more but these were the most important
stats_data = {
    "TOTAL URLs": 0,
    "Fetched pages (status 2)": 0,
    "Unfetched pages (status 1)": 0,
    "Redirects (status 5)": 0,
    "avg fetch time" : 0.0,
    "avg fetch interval" : 0.0,
    "avg score" : 0.0,
    "min score" : 0.0,
    "max score" : 0.0
}
#putting stats in the right place
for line in out.split("\n"):
    line = line.strip()
    if "TOTAL urls:" in line:
        stats_data["TOTAL URLs"] = int(line.split(":")[-1].strip())
    elif "status 1 (db_unfetched):" in line:
        stats_data["Unfetched pages (status 1)"] = int(line.split(":")[-1].strip())
    elif "status 2 (db_fetched):" in line:
        stats_data["Fetched pages (status 2)"] = int(line.split(":")[-1].strip())
    elif "status 5 (db_redir_perm):" in line:
        stats_data["Redirects (status 5)"] = int(line.split(":")[-1].strip())
    elif "avg of fetch times:" in line:
        val_str = line.split(":")[-1].strip()
        match = re.match(r"^([\d\.]+)", val_str)
        stats_data["avg fetch time"] = float(match.group(1)) if match else 0.0
    elif "avg fetch interval:" in line:
        val_str = line.split(":")[-1].strip()
        match = re.match(r"^([\d\.]+)", val_str)
        stats_data["avg fetch interval"] = float(match.group(1)) if match else 0.0
    elif "min score" in line:
        val_str = line.split(":")[-1].strip()
        match = re.match(r"^([\d\.]+)", val_str)
        stats_data["min score"] = float(match.group(1)) if match else 0.0
    elif "max score:" in line:
        val_str = line.split(":")[-1].strip()
        match = re.match(r"^([\d\.]+)", val_str)
        stats_data["max score"] = float(match.group(1)) if match else 0.0
    elif "avg score:" in line:
        val_str = line.split(":")[-1].strip()
        match = re.match(r"^([\d\.]+)", val_str)
        stats_data["avg score"] = float(match.group(1)) if match else 0.0

df_stats_cmd = pd.DataFrame(list(stats_data.items()), columns=["Metric Name", "Value"])

try:
    with open(LOG_PATH, "r", encoding="utf-8") as log_file:
        log_lines = [line.strip() for line in log_file if line.strip()]

    if log_lines:
        start_match = re.match(r"^(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})", log_lines[0])
        end_match = re.match(r"^(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})", log_lines[-1])

        if start_match and end_match:
            start_ts = pd.to_datetime(start_match.group(1), format="%Y-%m-%d %H:%M:%S")
            end_ts = pd.to_datetime(end_match.group(1), format="%Y-%m-%d %H:%M:%S")
            total_time_seconds = (end_ts - start_ts).total_seconds()
            total_time_minutes = total_time_seconds / 60.0
        else:
            total_time_minutes = 0
    else:
        total_time_minutes = 0
except Exception:
    total_time_minutes = 0

if total_time_minutes > 0:
    crawl_speed = stats_data["Fetched pages (status 2)"] / total_time_minutes
else:
    crawl_speed = 0.0

if stats_data["TOTAL URLs"] > 0:
    crawl_ratio = stats_data["Fetched pages (status 2)"] / stats_data["TOTAL URLs"]
else:
    crawl_ratio = 0

metrics = pd.DataFrame([
    { "Metric Name": "Crawl Speed (pages/minute)", "Value": round(crawl_speed, 2) },
    { "Metric Name": "Ratio (#URL crawled / #URL total)", "Value": round(crawl_ratio, 4) }
])

df_stats_cmd = pd.concat([df_stats_cmd, metrics], ignore_index=True)
keyword_dict, global_counter = get_keyword_counts(SEG_PATH)
df['URL_Clean'] = df['URL'].str.strip().str.rstrip('/').str.lower()
keyword_dict_lower = {k.lower(): v for k, v in keyword_dict.items()}

df_keywords = pd.DataFrame(list(keyword_dict_lower.items()), columns=['URL_Clean', 'Keywords_Found'])
df = pd.merge(df, df_keywords, on='URL_Clean', how='outer')

df['URL'] = df['URL'].fillna(df['URL_Clean'])
df['Keywords_Found'] = df['Keywords_Found'].fillna(0)
df['Status'] = df['Status'].fillna('2 (db_fetched)')
df['Fetch Time'] = df['Fetch Time'].fillna('Unknown (Segment Only)')

url_fetch_times = {}
try:
    with open(LOG_PATH, "r", encoding="utf-8") as log_file:
        for line in log_file:
            #basically check for website headers and take the tmime
            match = re.search(r"^(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}).*?fetching\s+(https?://\S+)", line, re.IGNORECASE)
            if match:
                ts = pd.to_datetime(match.group(1))
                url = match.group(2).strip().rstrip('/').lower()
                url_fetch_times[url] = ts
except FileNotFoundError:
    pass

df['Real_Fetch_Time'] = df['URL_Clean'].map(url_fetch_times)

#graphing, 4 graphs
def parse_dump_time(t_str):
    if pd.isna(t_str) or 'Unknown' in str(t_str):
        return pd.NaT
    try:
        clean_str = re.sub(r'\s+[A-Z]{3,4}\s+', ' ', str(t_str))
        return pd.to_datetime(clean_str)
    except:
        #make sure to take out
        return pd.NaT

df['Dump_Time_Parsed'] = df['Fetch Time'].apply(parse_dump_time)
df['Real_Fetch_Time'] = df['Real_Fetch_Time'].fillna(df['Dump_Time_Parsed'])
#fetch time has days, so use local run from hadoop
try:
    df['Real_Fetch_Time'] = df['Real_Fetch_Time'].fillna(start_ts)
except NameError:
    df['Real_Fetch_Time'] = df['Real_Fetch_Time'].bfill().ffill()

df_success = df[(df['Keywords_Found'] > 0) | (df['Status'].str.contains('2', na=False))].dropna(subset=['Real_Fetch_Time']).copy()

if not df_success.empty:
    df_success['Minute_Bin'] = df_success['Real_Fetch_Time'].dt.floor('min')

    timeline = df_success.groupby('Minute_Bin').agg({
        'URL': 'count',
        'Keywords_Found': 'sum'
    })

    full_time_range = pd.date_range(start=timeline.index.min(), end=timeline.index.max(), freq='min')
    timeline = timeline.reindex(full_time_range, fill_value=0)

    try:
        final_time = end_ts
        full_time_range = pd.date_range(start=timeline.index.min(), end=final_time, freq='min')
        timeline = timeline.reindex(full_time_range, fill_value=0)
    except NameError:
        pass

    timeline['Speed'] = timeline['URL'].rolling(window=3, min_periods=1).mean()
    timeline['Cumulative_Fetched'] = timeline['URL'].cumsum()

    total_urls = stats_data.get("TOTAL URLs", len(df))
    timeline['Unfetched_Queue'] = total_urls - timeline['Cumulative_Fetched']

    fig, (ax1, ax4, ax5, ax2, ax3) = plt.subplots(5, 1, figsize=(12, 35))

    #crawl speed, based on FETCHED pages, not unfetched
    ax1.plot(timeline.index, timeline['Speed'], color='blue', marker='o', linewidth=2)
    ax1.set_title('Successful Crawl Speed (fetched)')
    ax1.set_ylabel('Pages/Minute')
    ax1.grid(True)

    #keywords per time
    ax4.plot(timeline.index, timeline['Keywords_Found'], color='purple', marker='o', linewidth=2)
    ax4.set_title('Unique Keywords Extracted per Minute')
    ax4.set_ylabel('Keywords Found')
    ax4.grid(True, axis='y')

    time_formatter = mdates.DateFormatter('%H:%M')
    time_locator = mdates.MinuteLocator(interval=1)

    for ax in [ax1, ax4]:
        ax.xaxis.set_major_locator(time_locator)
        ax.xaxis.set_major_formatter(time_formatter)
        ax.set_xlim(timeline.index.min(), timeline.index.max())
        ax.set_xlabel('Time')

        for tick in ax.get_xticklabels():
            tick.set_rotation(45)

    #crawl speed vs keywords
    ax5.scatter(timeline['Speed'], timeline['Keywords_Found'], color='orange', s=60, alpha=0.7)
    ax5.set_title('Keywords Extracted vs. Crawl Speed')
    ax5.set_xlabel('Crawl Speed (Fetched Pages/Minute)')
    ax5.set_ylabel('Keywords Found')
    ax5.grid(True)

    #crawl speed vs fetches (should be basically the same to graph 1)
    ax2.scatter(timeline['Cumulative_Fetched'], timeline['Speed'], color='teal', s=50)
    ax2.set_title('Crawl Speed vs Total Fetch Sucesses')
    ax2.set_xlabel('Cumulative Successful Fetches')
    ax2.set_ylabel('Pages/Minute')
    ax2.grid(True)

    #fetched queue vs crawl speed
    ax3.scatter(timeline['Unfetched_Queue'], timeline['Speed'], color='red', s=50)
    ax3.set_title('Crawl Speed vs Unfetched Queue')
    ax3.set_xlabel('Unfetched URLs Remaining')
    ax3.set_ylabel('Pages/Minute')
    ax3.grid(True)

    plt.tight_layout()
    plt.savefig(OUTPUT_IMAGE)
    plt.close()

df_top_keywords = pd.DataFrame(global_counter.most_common(50), columns=["Keyword", "Frequency"])

#writing excel
with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Raws", index=False)
    df_stats_cmd.to_excel(writer, sheet_name="Analysis + stats", index=False)
    df_top_keywords.to_excel(writer, sheet_name="Top Keywords", index=False)

#checking spike reason
#df_anomaly = df_success[df_success['Minute_Bin'].dt.strftime('%H:%M') == '18:28']
#print("\nlooking at 18:28 spike-")
#print(df_anomaly['URL'].head(20).to_string())

wb = load_workbook(OUTPUT)
ws = wb.create_sheet("Graphs")
img = Image(OUTPUT_IMAGE)
ws.add_image(img, 'A1')
wb.save(OUTPUT)